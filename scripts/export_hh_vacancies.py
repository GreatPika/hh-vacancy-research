#!/usr/bin/env python3
"""Export hh-vacancy-research JSON results to JSON, Markdown, CSV, and XLSX."""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from collections import Counter
from pathlib import Path

# pyright: reportMissingModuleSource=false


EXCEL_CELL_LIMIT = 32767
DESCRIPTION_CHUNK_SIZE = 30000
LARGE_XLSX_ROW_THRESHOLD = 1000


def require_openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ModuleNotFoundError as exc:
        requirements = Path(__file__).resolve().parents[1] / "requirements.txt"
        raise SystemExit(
            "openpyxl is required for XLSX export. "
            f"Install dependencies with: pip install -r {requirements}"
        ) from exc
    return Workbook, WriteOnlyCell, Alignment, Font, PatternFill, Table, TableStyleInfo


def atomic_write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(f".{path.name}.{os.getpid()}.tmp")
    tmp_path.write_text(text, encoding="utf-8")
    tmp_path.replace(path)


def atomic_write_csv(path: Path, rows: list[list[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(f".{path.name}.{os.getpid()}.tmp")
    with tmp_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        for row in rows:
            writer.writerow([spreadsheet_safe_cell(value) for value in row])
    tmp_path.replace(path)


def atomic_write_markdown(path: Path, data: dict[str, object], rows: list[list[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(f".{path.name}.{os.getpid()}.tmp")
    with tmp_path.open("w", encoding="utf-8") as handle:
        handle.write(f"# {data.get('title') or 'hh.ru vacancies'}\n\n")
        handle.write(f"Vacancies: {len(rows) - 1}.\n\n")
        handle.write("| " + " | ".join(markdown_cell(value) for value in rows[0]) + " |\n")
        handle.write("| " + " | ".join("---" for _ in rows[0]) + " |\n")
        for row in rows[1:]:
            handle.write("| " + " | ".join(markdown_cell(value) for value in row) + " |\n")
    tmp_path.replace(path)


def spreadsheet_safe_cell(value: object) -> str:
    text = str(value or "")
    stripped = text.lstrip(" \t\r\n")
    if text.startswith(("\t", "\r", "\n")) or stripped.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def xlsx_safe_cell(value: object) -> str:
    text = spreadsheet_safe_cell(value)
    if len(text) <= EXCEL_CELL_LIMIT:
        return text
    return "[see Descriptions sheet: text exceeds Excel's per-cell limit]"


def needs_description_chunks(value: object) -> bool:
    return len(spreadsheet_safe_cell(value)) > EXCEL_CELL_LIMIT


def force_xlsx_string_cell(cell, value: object):
    cell.value = str(value or "")
    cell.data_type = "s"
    if hasattr(cell, "quotePrefix"):
        cell.quotePrefix = True
    return cell


def write_only_string_cell(sheet, value: object, WriteOnlyCell):
    cell = WriteOnlyCell(sheet, value=str(value or ""))
    cell.data_type = "s"
    if hasattr(cell, "quotePrefix"):
        cell.quotePrefix = True
    return cell


def markdown_cell(value: object) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", "<br>").strip()


def object_list(value: object) -> list[object]:
    return value if isinstance(value, list) else []


def object_dict(value: object) -> dict[str, object]:
    return value if isinstance(value, dict) else {}


def string_list(value: object) -> list[str]:
    return [item for item in object_list(value) if isinstance(item, str)]


def dict_list(value: object) -> list[dict[str, object]]:
    return [item for item in object_list(value) if isinstance(item, dict)]


def match_fields(vacancy: dict[str, object]) -> str:
    fields: list[str] = []
    for match in dict_list(vacancy.get("matches")):
        for field in object_list(match.get("fields")):
            if isinstance(field, str) and field not in fields:
                fields.append(field)
    return ", ".join(fields)


def matched_terms(vacancy: dict[str, object]) -> str:
    terms = vacancy.get("matched_terms")
    if isinstance(terms, list):
        return ", ".join(str(term) for term in terms)
    return ", ".join(
        str(match.get("term"))
        for match in dict_list(vacancy.get("matches"))
        if match.get("term")
    )


def structured_matched_terms(vacancy: dict[str, object]) -> list[str]:
    terms = vacancy.get("matched_terms")
    if isinstance(terms, list):
        return [term for term in terms if isinstance(term, str) and term]
    result: list[str] = []
    for match in dict_list(vacancy.get("matches")):
        term = match.get("term")
        if isinstance(term, str) and term:
            result.append(term)
    return result


def rows_for(vacancies: list[dict[str, object]]) -> list[list[str]]:
    rows = [
        [
            "Title",
            "Company",
            "URL",
            "Matched groups",
            "Matched fields",
            "Skills",
            "Description",
        ]
    ]
    for vacancy in sorted(vacancies, key=lambda item: str(item.get("title", "")).lower()):
        rows.append(
            [
                str(vacancy.get("title", "")),
                str(vacancy.get("company", "")),
                str(vacancy.get("url", "")),
                matched_terms(vacancy),
                match_fields(vacancy),
                ", ".join(str(skill) for skill in object_list(vacancy.get("skills"))),
                str(vacancy.get("description", "")),
            ]
        )
    return rows


def validate_input(data: object, source_path: Path) -> dict[str, object]:
    if not isinstance(data, dict):
        raise ValueError(f"{source_path}: root must be an object")
    vacancies = data.get("vacancies")
    if not isinstance(vacancies, list):
        raise ValueError(f"{source_path}: vacancies must be a list")
    for index, vacancy in enumerate(vacancies, start=1):
        if not isinstance(vacancy, dict):
            raise ValueError(f"{source_path}: vacancies[{index}] must be an object")
        for field in ("title", "company", "url", "description"):
            if field in vacancy and not isinstance(vacancy[field], str):
                raise ValueError(f"{source_path}: vacancies[{index}].{field} must be a string")
        if "skills" in vacancy and not isinstance(vacancy["skills"], list):
            raise ValueError(f"{source_path}: vacancies[{index}].skills must be a list")
        if "matches" in vacancy and not isinstance(vacancy["matches"], list):
            raise ValueError(f"{source_path}: vacancies[{index}].matches must be a list")
    return data


def load_json_object(path: Path) -> dict[str, object]:
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise ValueError(f"{path}: invalid JSON: {exc}") from exc
    return validate_input(raw, path)


def recompute_summary(data: dict[str, object]) -> dict[str, object]:
    vacancies = dict_list(data.get("vacancies"))
    counter: Counter[str] = Counter()
    for vacancy in vacancies:
        counter.update(structured_matched_terms(vacancy))
    summary = dict(object_dict(data.get("summary")))
    summary["kept"] = len(vacancies)
    summary["top_terms"] = dict(counter.most_common())
    return summary


def write_json(path: Path, data: dict[str, object]) -> None:
    output = dict(data)
    output["summary"] = recompute_summary(data)
    atomic_write_text(path, json.dumps(output, ensure_ascii=False, indent=2))


def write_markdown(path: Path, data: dict[str, object], rows: list[list[str]]) -> None:
    atomic_write_markdown(path, data, rows)


def write_csv(path: Path, rows: list[list[str]]) -> None:
    atomic_write_csv(path, rows)


def write_xlsx(path: Path, rows: list[list[str]]) -> None:
    Workbook, WriteOnlyCell, Alignment, Font, PatternFill, Table, TableStyleInfo = require_openpyxl()
    if len(rows) > LARGE_XLSX_ROW_THRESHOLD:
        write_xlsx_streaming(path, rows, Workbook, WriteOnlyCell)
        return

    workbook = Workbook()
    sheet = workbook.active
    if sheet is None:
        raise ValueError("workbook has no active worksheet")
    sheet.title = "Vacancies"
    for row in rows:
        sheet.append([xlsx_safe_cell(value) for value in row])

    if len(rows) > 1:
        table_range = f"A1:G{len(rows)}"
        table = Table(displayName="VacanciesTable", ref=table_range)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 36,
        "B": 28,
        "C": 28,
        "D": 30,
        "E": 20,
        "F": 42,
        "G": 100,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row_number = row[0].row
        if row_number is not None:
            sheet.row_dimensions[row_number].height = 90
    sheet.freeze_panes = "A2"
    add_description_chunks_sheet(workbook, rows, Alignment)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(f".{path.name}.{os.getpid()}.tmp")
    workbook.save(tmp_path)
    tmp_path.replace(path)


def write_xlsx_streaming(path: Path, rows: list[list[str]], Workbook, WriteOnlyCell) -> None:
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Vacancies")
    for row in rows:
        sheet.append([xlsx_safe_cell(value) for value in row])

    long_rows = (
        (index, row[0], row[2], row[6])
        for index, row in enumerate(rows[1:], start=1)
        if needs_description_chunks(row[6])
    )
    descriptions = None
    for vacancy_row, title, url, description in long_rows:
        text = str(description or "")
        for chunk_index, start in enumerate(range(0, len(text), DESCRIPTION_CHUNK_SIZE), start=1):
            if descriptions is None:
                descriptions = workbook.create_sheet("Descriptions")
                descriptions.append(["Vacancy row", "Title", "URL", "Chunk", "Description text"])
            descriptions.append([
                vacancy_row + 1,
                xlsx_safe_cell(title),
                xlsx_safe_cell(url),
                chunk_index,
                write_only_string_cell(
                    descriptions,
                    text[start : start + DESCRIPTION_CHUNK_SIZE],
                    WriteOnlyCell,
                ),
            ])

    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(f".{path.name}.{os.getpid()}.tmp")
    workbook.save(tmp_path)
    tmp_path.replace(path)


def add_description_chunks_sheet(workbook, rows: list[list[str]], Alignment) -> None:
    has_long_description = any(
        needs_description_chunks(row[6])
        for row in rows[1:]
    )
    if not has_long_description:
        return

    sheet = workbook.create_sheet("Descriptions")
    sheet.append(["Vacancy row", "Title", "URL", "Chunk", "Description text"])
    for vacancy_row, title, url, description in (
        (index, row[0], row[2], row[6])
        for index, row in enumerate(rows[1:], start=1)
        if needs_description_chunks(row[6])
    ):
        text = str(description or "")
        for chunk_index, start in enumerate(range(0, len(text), DESCRIPTION_CHUNK_SIZE), start=1):
            sheet.append([
                vacancy_row + 1,
                xlsx_safe_cell(title),
                xlsx_safe_cell(url),
                chunk_index,
                "",
            ])
            force_xlsx_string_cell(sheet.cell(row=sheet.max_row, column=5), text[start : start + DESCRIPTION_CHUNK_SIZE])
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 36
    sheet.column_dimensions["C"].width = 28
    sheet.column_dimensions["D"].width = 10
    sheet.column_dimensions["E"].width = 100
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export hh-vacancy-research JSON results to JSON, Markdown, CSV, and XLSX."
    )
    parser.add_argument("--source-json", type=Path, required=True, help="Source JSON produced by hh_vacancy_scraper.py.")
    parser.add_argument("--output-dir", type=Path, default=Path("."), help="Directory for exported files.")
    parser.add_argument("--output-prefix", default="", help="Output filename prefix. Defaults to '<source-stem>_export'.")
    return parser.parse_args()


def output_paths(args: argparse.Namespace) -> dict[str, Path]:
    prefix = args.output_prefix or f"{args.source_json.stem}_export"
    prefix_path = Path(prefix)
    if prefix_path.is_absolute() or len(prefix_path.parts) != 1 or prefix in {"", ".", ".."}:
        raise ValueError("--output-prefix must be a filename stem, not a path")
    output_dir = args.output_dir
    return {
        "json": output_dir / f"{prefix}.json",
        "md": output_dir / f"{prefix}.md",
        "csv": output_dir / f"{prefix}.csv",
        "xlsx": output_dir / f"{prefix}.xlsx",
    }


def ensure_source_not_overwritten(source_json: Path, outputs: dict[str, Path]) -> None:
    source = source_json.resolve()
    for kind, path in outputs.items():
        if path.resolve() == source:
            raise ValueError(
                f"{kind} output would overwrite source JSON. "
                "Use a different --output-dir or --output-prefix."
            )


def skill_root() -> Path:
    return Path(__file__).resolve().parents[1]


def is_relative_to(path: Path, parent: Path) -> bool:
    try:
        path.resolve().relative_to(parent.resolve())
    except ValueError:
        return False
    return True


def validate_work_paths(args: argparse.Namespace, outputs: dict[str, Path]) -> None:
    root = skill_root()
    if is_relative_to(args.output_dir, root):
        raise ValueError("--output-dir must not point inside the skill package")
    for label, path in {"output-dir": args.output_dir, **outputs}.items():
        if is_relative_to(path, root):
            raise ValueError(
                f"{label.upper()} output path would be inside the skill package; "
                "change --output-dir or --output-prefix"
            )


def main() -> int:
    args = parse_args()
    outputs = output_paths(args)
    ensure_source_not_overwritten(args.source_json, outputs)
    validate_work_paths(args, outputs)
    data = load_json_object(args.source_json)
    require_openpyxl()
    vacancies = dict_list(data.get("vacancies"))
    rows = rows_for(vacancies)

    write_json(outputs["json"], data)
    write_markdown(outputs["md"], data, rows)
    write_csv(outputs["csv"], rows)
    write_xlsx(outputs["xlsx"], rows)

    print(f"wrote {outputs['json']}")
    print(f"wrote {outputs['md']}")
    print(f"wrote {outputs['csv']}")
    print(f"wrote {outputs['xlsx']}")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except (OSError, json.JSONDecodeError, ValueError) as exc:
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(2)
